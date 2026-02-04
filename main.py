import asyncio
import logging
import json
import urllib.parse
import time
import os
from aiohttp import web
from openpyxl import Workbook, load_workbook
from io import BytesIO

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import WebAppInfo, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile

import database 

TOKEN = "8451254918:AAHDJ8yIwQ44Stn7vT_s1mmxLDVYwfXUuJU"
# Ссылка на сайт остается прежней
WEB_APP_URL = "https://rikman21.github.io/Gorbushka/?v=2"
# ID администратора (твой ID)
ADMIN_ID = 464896073 

logging.basicConfig(level=logging.INFO)
bot = Bot(token=TOKEN)
dp = Dispatcher()

# --- API СЕРВЕР (ТЕПЕРЬ УМНЫЙ) ---
async def health_check(request):
    return web.Response(text="Alive")

async def get_offers_api(request):
    # Эта функция отправляет цены сайту, когда он просит
    offers = database.get_all_offers_for_web()
    return web.json_response(offers, headers={
        "Access-Control-Allow-Origin": "*",  # Разрешаем доступ с GitHub Pages
        "Access-Control-Allow-Methods": "GET, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type"
    })

async def start_server():
    port = int(os.environ.get("PORT", 8080))
    app = web.Application()
    # Два маршрута: проверка жизни и выдача товаров
    app.router.add_get('/', health_check)
    app.router.add_get('/api/offers', get_offers_api)
    
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, '0.0.0.0', port)
    await site.start()

# --- EXCEL ---
def generate_excel_template(category_filter=None):
    products = database.get_catalog_for_excel(category_filter)
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"
    headers = ["SKU (Не менять!)", "Модель", "Память", "Цвет", "Сим", "ВАША ЦЕНА (Рубли)"]
    ws.append(headers)
    for p in products:
        row = list(p) + [""] 
        ws.append(row)
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.read()

# --- HANDLERS ---
@dp.message(F.document)
async def handle_document(message: types.Message):
    if not message.document.file_name.endswith('.xlsx'):
        return await message.answer("❌ Это не Excel. Пришлите файл .xlsx")

    user_id = message.from_user.id
    username = message.from_user.username or "Продавец"
    wait_msg = await message.answer("⏳ Обрабатываю прайс...")

    try:
        bot_file = await bot.get_file(message.document.file_id)
        file_data = await bot.download_file(bot_file.file_path)
        wb = load_workbook(file_data)
        ws = wb.active
        prices_to_update = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Ожидаемый формат строки Excel:
            # [0] SKU, [1] Модель, [2] Память, [3] Цвет, [4] SIM, [5] Цена
            sku = row[0]
            model = row[1] if len(row) > 1 else None
            memory = row[2] if len(row) > 2 else None
            color = row[3] if len(row) > 3 else None
            sim_type = row[4] if len(row) > 4 else None
            price_raw = row[5] if len(row) > 5 else None

            price = None
            if price_raw is not None and price_raw != "":
                try:
                    price = int(str(price_raw).replace(" ", "").replace("₽", ""))
                except Exception:
                    price = None

            if sku and price is not None:
                prices_to_update.append((sku, model, memory, color, sim_type, price))
        
        updated_count, skipped_count = database.update_prices_from_excel(user_id, username, prices_to_update)
        
        msg = f"✅ **Прайс обновлен!**\n\nТоваров в продаже: {updated_count}"
        if skipped_count > 0:
            msg += f"\n⚠️ Пропущено невалидных SKU: {skipped_count}"
        msg += "\n\nТеперь просто откройте WebApp, ссылка обновлять не нужно."
        
        await wait_msg.edit_text(msg)
        
    except Exception as e:
        logging.error(e)
        await wait_msg.edit_text("❌ Ошибка. Проверьте формат файла.")

@dp.message(Command("start"))
async def start(message: types.Message):
    user_id = message.from_user.id
    
    # БОЛЬШЕ НЕ ПЕРЕДАЕМ ДАННЫЕ В ССЫЛКЕ
    # Ссылка теперь короткая и вечная
    full_url = f"{WEB_APP_URL}?uid={user_id}"

    kb = [[KeyboardButton(text="📱 ОТКРЫТЬ МАРКЕТ", web_app=WebAppInfo(url=full_url))]]
    await message.answer("👋 Горбушка Онлайн v2.0 (Live)", reply_markup=ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True))

@dp.message(Command("admin"))
async def admin_command(message: types.Message):
    user_id = message.from_user.id
    if user_id != ADMIN_ID:
        await message.answer("❌ У вас нет прав доступа к этой команде.")
        return
    
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Добавить товар", callback_data="admin_add")],
        [InlineKeyboardButton(text="🗑 Удалить товар", callback_data="admin_delete")],
        [InlineKeyboardButton(text="🔍 Найти товар", callback_data="admin_search")],
        [InlineKeyboardButton(text="📋 Категории", callback_data="admin_categories")]
    ])
    await message.answer("🔧 **Панель администратора**\n\nВыберите действие:", reply_markup=kb, parse_mode="HTML")

@dp.callback_query(F.data == "admin_categories")
async def admin_categories(callback: types.CallbackQuery):
    categories = database.get_catalog_categories()
    kb_buttons = []
    for cat in categories:
        kb_buttons.append([InlineKeyboardButton(text=f"📥 Скачать шаблон ({cat})", callback_data=f"template_{cat}")])
    kb_buttons.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="admin_back")])
    kb = InlineKeyboardMarkup(inline_keyboard=kb_buttons)
    await callback.message.edit_text(f"📋 **Категории каталога:**\n\n" + "\n".join([f"• {cat}" for cat in categories]), reply_markup=kb, parse_mode="HTML")

@dp.callback_query(F.data.startswith("template_"))
async def admin_template_category(callback: types.CallbackQuery):
    category = callback.data.split("_", 1)[1]
    file_bytes = generate_excel_template(category)
    filename = f"Gorbushka_Price_Template_{category}.xlsx"
    document = BufferedInputFile(file_bytes, filename=filename)
    await callback.message.answer_document(document, caption=f"📉 **Шаблон для цен ({category})**")
    await callback.answer(f"Шаблон для {category} отправлен")

@dp.callback_query(F.data == "admin_add")
async def admin_add(callback: types.CallbackQuery):
    await callback.message.edit_text(
        "➕ **Добавление товара в каталог**\n\n"
        "Отправьте данные в формате:\n"
        "`/add SKU|Модель|Память|Цвет|SIM-тип|Категория`\n\n"
        "Пример:\n"
        "`/add iPhone16PM_256_Black_Dual|iPhone 16 Pro Max|256GB|Черный титан|Dual|iPhone`",
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.message(Command("add"))
async def add_product(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        return
    
    try:
        parts = message.text.split(" ", 1)[1].split("|")
        if len(parts) != 6:
            await message.answer("❌ Неверный формат. Используйте:\n`/add SKU|Модель|Память|Цвет|SIM-тип|Категория`")
            return
        
        sku, model, memory, color, sim_type, category = [p.strip() for p in parts]
        success, msg = database.add_product_to_catalog(sku, model, memory, color, sim_type, category)
        
        if success:
            await message.answer(f"✅ {msg}\n\nSKU: `{sku}`", parse_mode="Markdown")
        else:
            await message.answer(f"❌ {msg}")
    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)}")

@dp.callback_query(F.data == "admin_delete")
async def admin_delete(callback: types.CallbackQuery):
    await callback.message.edit_text(
        "🗑 **Удаление товара из каталога**\n\n"
        "Отправьте команду:\n"
        "`/delete SKU`\n\n"
        "Пример:\n"
        "`/delete iPhone16PM_256_Black_Dual`",
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.message(Command("delete"))
async def delete_product(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        return
    
    try:
        sku = message.text.split(" ", 1)[1].strip()
        success, msg = database.delete_product_from_catalog(sku)
        
        if success:
            await message.answer(f"✅ {msg}\n\nSKU: `{sku}`", parse_mode="Markdown")
        else:
            await message.answer(f"❌ {msg}")
    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)}")

@dp.callback_query(F.data == "admin_search")
async def admin_search(callback: types.CallbackQuery):
    await callback.message.edit_text(
        "🔍 **Поиск товара в каталоге**\n\n"
        "Отправьте команду:\n"
        "`/search запрос`\n\n"
        "Пример:\n"
        "`/search iPhone 16`",
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.message(Command("search"))
async def search_product(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        return
    
    try:
        query = message.text.split(" ", 1)[1].strip()
        results = database.search_catalog(query)
        
        if not results:
            await message.answer(f"❌ Товары не найдены по запросу: `{query}`", parse_mode="Markdown")
            return
        
        text = f"🔍 **Найдено товаров: {len(results)}**\n\n"
        for sku, model, memory, color, sim_type, category in results[:10]:
            text += f"• `{sku}`\n  {model} {memory or ''} {color or ''} [{sim_type or ''}] ({category})\n\n"
        
        if len(results) > 10:
            text += f"... и еще {len(results) - 10} товаров"
        
        await message.answer(text, parse_mode="Markdown")
    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)}")

@dp.callback_query(F.data == "admin_back")
async def admin_back(callback: types.CallbackQuery):
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Добавить товар", callback_data="admin_add")],
        [InlineKeyboardButton(text="🗑 Удалить товар", callback_data="admin_delete")],
        [InlineKeyboardButton(text="🔍 Найти товар", callback_data="admin_search")],
        [InlineKeyboardButton(text="📋 Категории", callback_data="admin_categories")]
    ])
    await callback.message.edit_text("🔧 **Панель администратора**\n\nВыберите действие:", reply_markup=kb, parse_mode="HTML")

@dp.message(F.web_app_data)
async def handle_webapp(message: types.Message):
    data = message.web_app_data.data
    user_id = message.chat.id
    username = message.from_user.username or "Клиент"

    if data == "REQ_TEMPLATE" or data.startswith("REQ_TEMPLATE|"):
        # Поддержка фильтрации по категории: REQ_TEMPLATE|iPhone
        category_filter = None
        if "|" in data:
            category_filter = data.split("|")[1] if len(data.split("|")) > 1 else None
            # Валидация категории
            valid_categories = database.get_catalog_categories()
            if category_filter not in valid_categories:
                category_filter = None
        
        file_bytes = generate_excel_template(category_filter)
        filename = f"Gorbushka_Price_Template_{category_filter}.xlsx" if category_filter else "Gorbushka_Price_Template.xlsx"
        caption = f"📉 **Шаблон для цен**" + (f" ({category_filter})" if category_filter else "")
        document = BufferedInputFile(file_bytes, filename=filename)
        await message.answer_document(document, caption=caption)
        return

    if data.startswith("DELETE_OFFER"):
        sku = data.split("|")[1]
        database.delete_offer_by_sku(user_id, sku)
        # Просто подтверждаем, обновлять ссылку не надо
        return

    if data.startswith("REQ_BUY"):
        parts = data.split("|")
        seller_id = int(parts[1])
        product_name = parts[3]
        price = parts[4]
        
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ В наличии", callback_data=f"confirm_{user_id}")],
            [InlineKeyboardButton(text="❌ Нет", callback_data=f"reject_{seller_id}")]
        ])
        try:
            await bot.send_message(seller_id, f"🔔 <b>ЗАКАЗ!</b>\n\n📦 {product_name}\n💰 {price}р\n👤 @{username}", reply_markup=kb, parse_mode="HTML")
            await message.answer("⏳ Запрос отправлен продавцу...")
        except:
            await message.answer("Продавец не найден.")

    elif data.startswith("NEW_PRICE"):
        # Формат: NEW_PRICE|Имя товара|Новая цена
        parts = data.split("|")
        if len(parts) < 3:
            await message.answer("Не удалось разобрать данные для изменения цены.")
            return

        product_name = parts[1]
        price_raw = parts[2]

        try:
            new_price = int(str(price_raw).replace(" ", "").replace("₽", ""))
        except ValueError:
            await message.answer("Цена указана неверно.")
            return

        affected = database.update_price_from_web(user_id, product_name, new_price)
        if affected > 0:
            await message.answer(f"✅ Цена для «{product_name}» обновлена на {new_price} ₽.")
        else:
            await message.answer("Товар не найден в вашей базе. Обновите прайс через Excel, а затем попробуйте снова.")

@dp.callback_query(F.data.startswith("confirm_"))
async def confirm_order(callback: types.CallbackQuery):
    buyer_id = int(callback.data.split("_")[1])
    await callback.message.edit_text(f"✅ Подтверждено!", reply_markup=None)
    await bot.send_message(buyer_id, f"🎉 Продавец подтвердил!\nКонтакт: @{callback.from_user.username}")

@dp.callback_query(F.data.startswith("reject_"))
async def reject_order(callback: types.CallbackQuery):
    await callback.message.edit_text(f"🚫 Отказ.", reply_markup=None)

async def main():
    database.init_db()
    await start_server()
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())

